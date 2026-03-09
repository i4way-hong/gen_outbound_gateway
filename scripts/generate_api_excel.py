import base64
import json
import os
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

from Crypto.Cipher import AES
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

BASE_DIR = Path(__file__).resolve().parents[1]
OPENAPI_PATH = Path(os.environ.get("OPENAPI_JSON", BASE_DIR / "docs" / "works" / "api-docs-2026-02-10.json"))
OUTPUT_PATH = Path(os.environ.get("OUTPUT_XLSX", BASE_DIR / "docs" / "works" / "api-spec-2026-02-10.xlsx"))
ENC_KEY = os.environ.get("CCC_SERVICE_ENC_KEY", "")
ENC_IV = os.environ.get("CCC_SERVICE_ENC_IV", "")

ENCRYPTED_PREFIXES = (
    "/api/v1/configuration",
    "/api/v1/stat",
    "/api/v1/tserver",
    "/api/v1/outbound",
    "/api/v1/crypto/secure",
)


def normalize_sheet_name(name: str, index: int) -> str:
    safe = "".join(ch for ch in name if ch not in "[]:*?/\\")
    safe = safe.strip().replace(" ", "_")
    if len(safe) > 25:
        safe = safe[:25]
    return f"{index:03d}_{safe}" if safe else f"API_{index:03d}"


def pretty_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def to_json_string(value: Any) -> str:
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("{") or text.startswith("["):
            return text
        return json.dumps(text, ensure_ascii=False)
    return json.dumps(value, ensure_ascii=False)


def format_example(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("{") or text.startswith("["):
            try:
                return pretty_json(json.loads(text))
            except json.JSONDecodeError:
                return text
        return text
    return pretty_json(value)


def parse_generic_block(text: str, start: int) -> Tuple[str, int]:
    depth = 0
    for idx in range(start, len(text)):
        ch = text[idx]
        if ch == "<":
            depth += 1
        elif ch == ">":
            depth -= 1
            if depth == 0:
                return text[start + 1: idx], idx + 1
    return "", start


def parse_api_response_types(source: str) -> Dict[str, str]:
    result: Dict[str, str] = {}
    idx = 0
    while True:
        pos = source.find("ApiResponse<", idx)
        if pos == -1:
            break
        inner, next_idx = parse_generic_block(source, pos + len("ApiResponse"))
        if not inner:
            idx = pos + 1
            continue
        cursor = next_idx
        while cursor < len(source) and source[cursor].isspace():
            cursor += 1
        name_start = cursor
        while cursor < len(source) and (source[cursor].isalnum() or source[cursor] == "_"):
            cursor += 1
        method_name = source[name_start:cursor]
        if method_name and cursor < len(source) and "(" in source[cursor:cursor + 3]:
            result[method_name] = inner.strip()
        idx = next_idx
    return result


def load_response_type_map(base_dir: Path) -> Dict[str, str]:
    type_map: Dict[str, str] = {}
    src_root = base_dir / "src" / "main" / "java"
    if not src_root.exists():
        return type_map
    for java_file in src_root.rglob("*.java"):
        try:
            text = java_file.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = java_file.read_text(encoding="cp949", errors="ignore")
        text = " ".join(text.split())
        type_map.update(parse_api_response_types(text))
    return type_map


def strip_package(type_name: str) -> str:
    return type_name.split(".")[-1]


def parse_type(type_name: str) -> Tuple[str, Optional[str]]:
    normalized = "".join(type_name.split())
    normalized = normalized.replace("?extends", "").replace("?super", "")
    if normalized.startswith("List<") or normalized.startswith("Set<"):
        inner, _ = parse_generic_block(normalized, normalized.find("<"))
        return "list", inner
    if normalized.startswith("Map<"):
        return "map", None
    if normalized in {"Void", "void"}:
        return "void", None
    return "object", strip_package(normalized)


def build_example(schema: Dict[str, Any], schemas: Dict[str, Any], depth: int = 0, visited: Optional[set] = None) -> Any:
    if visited is None:
        visited = set()
    if depth > 6:
        return None
    if not schema:
        return None
    if "$ref" in schema:
        ref_name = schema["$ref"].split("/")[-1]
        if ref_name in visited:
            return None
        visited.add(ref_name)
        return build_example(schemas.get(ref_name, {}), schemas, depth + 1, visited)
    if "example" in schema:
        return schema["example"]
    if "enum" in schema and schema["enum"]:
        return schema["enum"][0]
    schema_type = schema.get("type")
    if schema_type == "array":
        item_example = build_example(schema.get("items", {}), schemas, depth + 1, visited)
        return [item_example]
    if schema_type == "object" or "properties" in schema:
        props = schema.get("properties", {})
        result: Dict[str, Any] = {}
        for key, value in props.items():
            result[key] = build_example(value, schemas, depth + 1, visited)
        if not result and "additionalProperties" in schema:
            result["key"] = build_example(schema.get("additionalProperties", {}), schemas, depth + 1, visited)
        return result
    if schema_type == "string":
        return "string"
    if schema_type == "integer":
        return 0
    if schema_type == "number":
        return 0
    if schema_type == "boolean":
        return True
    return None


def build_data_example(type_name: str, schemas: Dict[str, Any]) -> Any:
    kind, inner = parse_type(type_name)
    if kind == "void":
        return None
    if kind == "map":
        return {"key": "value"}
    if kind == "list":
        if not inner:
            return []
        item_kind, item_name = parse_type(inner)
        if item_kind == "object" and item_name:
            return [build_example({"$ref": f"#/components/schemas/{item_name}"}, schemas)]
        if item_kind == "map":
            return [{"key": "value"}]
        return []
    if inner:
        return build_example({"$ref": f"#/components/schemas/{inner}"}, schemas)
    return None


def build_api_response_example(message: str, data_example: Any) -> Dict[str, Any]:
    return {
        "success": True,
        "message": message or "요청 성공",
        "data": data_example,
        "timestamp": "2026-03-04T09:00:00+09:00",
    }


def has_empty_data(example_text: str) -> bool:
    if not example_text:
        return True
    try:
        payload = json.loads(example_text)
    except json.JSONDecodeError:
        return False
    if not isinstance(payload, dict):
        return False
    if "data" not in payload:
        return False
    data_value = payload.get("data")
    return data_value is None or data_value == {} or data_value == []


def pad_pkcs7(raw: bytes, block_size: int = 16) -> bytes:
    padding = block_size - (len(raw) % block_size)
    return raw + bytes([padding]) * padding


def encrypt_encdata(plain_text: str) -> str:
    if not ENC_KEY or not ENC_IV:
        return ""
    cipher = AES.new(ENC_KEY.encode("utf-8"), AES.MODE_CBC, ENC_IV.encode("utf-8"))
    encrypted = cipher.encrypt(pad_pkcs7(plain_text.encode("utf-8")))
    return base64.b64encode(encrypted).decode("utf-8")


def extract_examples(content: Dict[str, Any]) -> Tuple[str, str, str]:
    if not content:
        return "", "", ""
    media_type = content.get("application/json") or next(iter(content.values()), {})
    examples = []
    if "examples" in media_type:
        for name, payload in media_type["examples"].items():
            examples.append((name, payload.get("value")))
    elif "example" in media_type:
        examples.append(("example", media_type.get("example")))

    plain = ""
    encrypted = ""
    plain_raw = ""
    for name, value in examples:
        if value is None:
            continue
        is_encrypted = False
        if name and "enc" in name.lower():
            is_encrypted = True
        if isinstance(value, dict) and set(value.keys()) == {"encData"}:
            is_encrypted = True
        if is_encrypted and not encrypted:
            encrypted = format_example(value)
        if not is_encrypted and not plain:
            plain = format_example(value)
            plain_raw = to_json_string(value)
    return plain, encrypted, plain_raw


def main() -> None:
    if not OPENAPI_PATH.exists():
        raise FileNotFoundError(f"OpenAPI 파일을 찾을 수 없습니다: {OPENAPI_PATH}")

    spec = json.loads(OPENAPI_PATH.read_text(encoding="utf-8"))
    paths = spec.get("paths", {})
    schemas = (spec.get("components") or {}).get("schemas", {})
    response_type_map = load_response_type_map(BASE_DIR)

    wb = Workbook()
    wb.remove(wb.active)

    list_sheet = wb.create_sheet("API_List")
    list_sheet.append([
        "No",
        "Sheet",
        "Method",
        "Path",
        "Summary",
        "OperationId",
        "Tags",
        "PrimaryTag",
        "Encrypted",
    ])

    header_font = Font(bold=True)
    for cell in list_sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    entries: List[Dict[str, Any]] = []
    for path, methods in paths.items():
        for method, operation in methods.items():
            if method.startswith("x-"):
                continue
            tag_list = operation.get("tags", [])
            primary_tag = tag_list[0] if tag_list else "Untagged"
            entries.append({
                "path": path,
                "method": method,
                "operation": operation,
                "summary": operation.get("summary", ""),
                "operation_id": operation.get("operationId", ""),
                "tags": ", ".join(tag_list),
                "primary_tag": primary_tag,
                "encrypted": "Y" if path.startswith(ENCRYPTED_PREFIXES) else "N",
            })

    entries.sort(key=lambda item: (item["primary_tag"], item["path"], item["method"]))

    api_index = 1
    group_map: Dict[str, List[Dict[str, Any]]] = {}
    for item in entries:
        group_map.setdefault(item["primary_tag"], []).append(item)

        sheet_name = normalize_sheet_name(f"{item['method']}_{item['path']}", api_index)
        list_sheet.append([
            api_index,
            sheet_name,
            item["method"].upper(),
            item["path"],
            item["summary"],
            item["operation_id"],
            item["tags"],
            item["primary_tag"],
            item["encrypted"],
        ])

        detail = wb.create_sheet(sheet_name)
        detail.append(["항목", "값"])
        detail["A1"].font = header_font
        detail["B1"].font = header_font

        detail.append(["Method", item["method"].upper()])
        detail.append(["Path", item["path"]])
        detail.append(["Summary", item["summary"]])
        detail.append(["Description", item["operation"].get("description", "")])
        detail.append(["Tags", item["tags"]])
        detail.append(["Primary Tag", item["primary_tag"]])
        detail.append(["Encrypted", item["encrypted"]])

        request_body = item["operation"].get("requestBody", {}) or {}
        request_content = request_body.get("content", {})
        req_plain, req_enc, req_plain_raw = extract_examples(request_content)
        if not req_plain:
            req_media = request_content.get("application/json") or next(iter(request_content.values()), {})
            req_schema = req_media.get("schema") if isinstance(req_media, dict) else None
            if req_schema:
                req_example = build_example(req_schema, schemas)
                if req_example is not None:
                    req_plain = format_example(req_example)
                    req_plain_raw = to_json_string(req_example)
        if item["encrypted"] == "Y" and not req_enc and req_plain_raw:
            enc_data = encrypt_encdata(req_plain_raw)
            if enc_data:
                req_enc = pretty_json({"encData": enc_data})
        detail.append(["Request Example (Plain)", req_plain])
        detail.append(["Request Example (Encrypted)", req_enc])

        response = (item["operation"].get("responses", {}) or {}).get("200", {})
        response_content = response.get("content", {}) if response else {}
        res_plain, res_enc, res_plain_raw = extract_examples(response_content)
        if has_empty_data(res_plain):
            op_id = item["operation"].get("operationId", "")
            response_type = response_type_map.get(op_id)
            if response_type:
                data_example = build_data_example(response_type, schemas)
                if data_example is not None:
                    api_example = build_api_response_example(item["summary"], data_example)
                    res_plain = format_example(api_example)
                    res_plain_raw = to_json_string(api_example)
        if item["encrypted"] == "Y" and not res_enc and res_plain_raw:
            enc_data = encrypt_encdata(res_plain_raw)
            if enc_data:
                res_enc = pretty_json({"encData": enc_data})
        detail.append(["Response Example (Plain)", res_plain])
        detail.append(["Response Example (Encrypted)", res_enc])

        for row in detail.iter_rows(min_row=1, max_row=detail.max_row, min_col=1, max_col=2):
            row[0].alignment = Alignment(vertical="top")
            row[1].alignment = Alignment(vertical="top", wrap_text=True)

        detail.column_dimensions["A"].width = 28
        detail.column_dimensions["B"].width = 100

        api_index += 1

    group_sheet = wb.create_sheet("API_Groups")
    group_sheet.append(["Group", "Count", "Sheet"])
    for cell in group_sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for index, (group, items) in enumerate(sorted(group_map.items()), start=1):
        group_sheet_name = normalize_sheet_name(f"Group_{group}", index)
        group_sheet.append([group, len(items), group_sheet_name])

        group_detail = wb.create_sheet(group_sheet_name)
        group_detail.append(["No", "Method", "Path", "Summary", "Encrypted"])
        for cell in group_detail[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_index, item in enumerate(items, start=1):
            group_detail.append([
                row_index,
                item["method"].upper(),
                item["path"],
                item["summary"],
                item["encrypted"],
            ])

        group_detail.column_dimensions["A"].width = 6
        group_detail.column_dimensions["B"].width = 10
        group_detail.column_dimensions["C"].width = 40
        group_detail.column_dimensions["D"].width = 50
        group_detail.column_dimensions["E"].width = 10

    list_sheet.column_dimensions["A"].width = 6
    list_sheet.column_dimensions["B"].width = 18
    list_sheet.column_dimensions["C"].width = 10
    list_sheet.column_dimensions["D"].width = 40
    list_sheet.column_dimensions["E"].width = 40
    list_sheet.column_dimensions["F"].width = 30
    list_sheet.column_dimensions["G"].width = 30
    list_sheet.column_dimensions["H"].width = 20
    list_sheet.column_dimensions["I"].width = 10

    group_sheet.column_dimensions["A"].width = 30
    group_sheet.column_dimensions["B"].width = 10
    group_sheet.column_dimensions["C"].width = 25

    wb.save(OUTPUT_PATH)
    print(f"Excel 생성 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
